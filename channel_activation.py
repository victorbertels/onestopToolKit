"""Collect channel activation data and build partner email templates."""

from __future__ import annotations

import csv
from datetime import date
from io import BytesIO, StringIO
from typing import Optional

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from utils import _channel_link_ids_from_location, getAllChannelLinks, getAllLocations

PARTNER_CHANNEL_IDS = {
    "Uber Eats": {7, 6007},
    "Deliveroo": {2, 6002},
    "Just Eat": {9, 6009},
}

PARTNER_ORDER = ("Uber Eats", "Deliveroo", "Just Eat")

# ---------------------------------------------------------------------------
# Email templates — edit the text below directly.
# Placeholders: {cohort}  {action_date}  {go_live_date}
# ---------------------------------------------------------------------------

SUBJECT_UBER_EATS = "Uber Eats activation — {cohort} — action by {action_date}"

INTRO_UBER_EATS = """\
We want to activate Uber Eats for the following sites.
Please complete the steps below so stores can go live on {go_live_date}."""

STEPS_UBER_EATS = """\
1. Link the following sites to feed from INCA.
2. Stage the following sites to Flooid Retail (Quest) client ID: jHZkWI0jkna2B_ksR825ARRJOU5Vwxvc
3. Confirm when the INCA linking is done so we can send all the files to the SFTP."""

SUBJECT_DELIVEROO = "Deliveroo activation — {cohort} — action on {action_date}"

INTRO_DELIVEROO = """\
We want to activate Deliveroo for the following sites.
Use the channel link IDs below to connect each site on {go_live_date}."""

STEPS_DELIVEROO = """\
1. Please link all the sites below to Quest enabled brand ID : deliverect-onestop-menu-questpicking 
2. Once that is done please let us know so we can update the listing for those sites.
"""

SUBJECT_JUST_EAT = "Just Eat activation — {cohort} — action on {action_date}"

INTRO_JUST_EAT = """\
We want to activate Just Eat for the following sites.
Please complete setup for each restaurant before {go_live_date}."""

STEPS_JUST_EAT = """\
1. By {action_date}, link each Just Eat restaurant below using the Deliverect channel link ID.

2. Please configure them for Quest:

Quest setup.

Path to send notifications for cancelled orders:
/flyt-retail/order/cancel-order-notification

Path to send failed orders:
/flyt-retail/order/cancel-order-notification

Path to process orders:
/flyt-retail/order

Path to process Final Picked Order:
/flyt-retail/order/final

Path to notify you when something has gone wrong with amending the order on the partner platform:
/flyt-retail/order/amendment-status-update

Path to notify restaurant that it has been temporarily set offline:
/flyt-retail/store-status-update

Path to notify restaurant of driver status updates:
/flyt-retail/order/driver-status-update

Final Picked Base URL:
https://api.deliverect.io

Base URL:
https://api.deliverect.io

3. Reply confirming activation is complete for all restaurants listed so we can publish the menus."""

PARTNER_EMAIL_TEMPLATES = {
    "Uber Eats": {
        "subject": SUBJECT_UBER_EATS,
        "intro": INTRO_UBER_EATS,
        "steps": STEPS_UBER_EATS,
    },
    "Deliveroo": {
        "subject": SUBJECT_DELIVEROO,
        "intro": INTRO_DELIVEROO,
        "steps": STEPS_DELIVEROO,
    },
    "Just Eat": {
        "subject": SUBJECT_JUST_EAT,
        "intro": INTRO_JUST_EAT,
        "steps": STEPS_JUST_EAT,
    },
}


def _fill_template(template: str, *, cohort: str, action_date: str, go_live_date: str) -> str:
    return template.format(
        cohort=cohort,
        action_date=action_date,
        go_live_date=go_live_date,
    )


def _channel_link_by_id(channel_links: list) -> dict:
    return {link["_id"]: link for link in channel_links if link.get("_id")}


def extract_unique_location_tags(locations: list) -> list[str]:
    tags = set()
    for location in locations:
        for tag in location.get("tags") or []:
            if tag:
                tags.add(tag)
    return sorted(tags, key=str.lower)


def filter_locations_by_tag(locations: list, tag: str) -> list:
    if not tag:
        return locations
    return [
        location
        for location in locations
        if tag in (location.get("tags") or [])
    ]


def _partner_for_channel(channel: int) -> Optional[str]:
    for partner, ids in PARTNER_CHANNEL_IDS.items():
        if channel in ids:
            return partner
    return None


def _partner_reference(partner: str, channel_link: dict) -> str:
    settings = channel_link.get("channelSettings") or {}
    if partner == "Uber Eats":
        return (settings.get("storeId") or "").strip()
    return channel_link.get("_id") or ""


def collect_activation_data(
    locations: list,
    channel_links: list,
    tag: str = "",
) -> dict[str, list[dict]]:
    """Group activation rows by delivery partner for locations matching tag."""
    filtered = filter_locations_by_tag(locations, tag) if tag else locations
    links_by_id = _channel_link_by_id(channel_links)
    grouped = {partner: [] for partner in PARTNER_ORDER}

    for location in filtered:
        location_name = location.get("name") or location.get("_id") or "Unknown location"
        location_id = location.get("_id") or ""

        for link_id in _channel_link_ids_from_location(location):
            channel_link = links_by_id.get(link_id)
            if not channel_link:
                continue

            partner = _partner_for_channel(channel_link.get("channel"))
            if not partner:
                continue

            grouped[partner].append(
                {
                    "location_name": location_name,
                    "location_id": location_id,
                    "channel_link_id": link_id,
                    "channel_link_name": channel_link.get("name") or "",
                    "partner_ref": _partner_reference(partner, channel_link),
                }
            )

    for partner in grouped:
        grouped[partner].sort(key=lambda row: row["location_name"].lower())

    return grouped


def fetch_activation_data(account: str, tag: str = "") -> dict[str, list[dict]]:
    locations = getAllLocations(account)
    channel_links = getAllChannelLinks(account)
    return collect_activation_data(locations, channel_links, tag=tag)


def _excel_sheet_headers(partner: str) -> list[str]:
    if partner == "Uber Eats":
        return [
            "Store name",
            "Uber Store ID",
            "Channel link ID",
            "Channel link name",
            "Location ID",
        ]
    return ["Store name", "Channel link ID", "Channel link name", "Location ID"]


def _excel_sheet_row(partner: str, row: dict) -> list:
    if partner == "Uber Eats":
        return [
            row["location_name"],
            row["partner_ref"] or "",
            row["channel_link_id"],
            row["channel_link_name"],
            row["location_id"],
        ]
    return [
        row["location_name"],
        row["channel_link_id"],
        row["channel_link_name"],
        row["location_id"],
    ]


def _autosize_worksheet_columns(ws) -> None:
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_length = 0
        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 60)


def activation_channels_excel_bytes(grouped: dict[str, list[dict]]) -> bytes:
    """Build an Excel workbook with one sheet per delivery partner."""
    wb = Workbook()
    wb.remove(wb.active)

    for partner in PARTNER_ORDER:
        ws = wb.create_sheet(title=partner[:31])
        headers = _excel_sheet_headers(partner)
        ws.append(headers)
        for row in grouped.get(partner) or []:
            ws.append(_excel_sheet_row(partner, row))
        _autosize_worksheet_columns(ws)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def activation_channels_excel_filename(cohort_tag: str) -> str:
    return f"channel_activation_{_cohort_slug(cohort_tag)}.xlsx"


def _cohort_slug(cohort_tag: str) -> str:
    slug = cohort_tag.strip().lower().replace(" ", "_").replace("/", "-")
    return "".join(ch for ch in slug if ch.isalnum() or ch in "-_") or "batch"


def _partner_csv_headers(partner: str) -> list[str]:
    if partner == "Uber Eats":
        return ["Store name", "Uber Store ID", "Channel link ID"]
    return ["Store name", "Channel link ID"]


def _partner_csv_row(partner: str, row: dict) -> list[str]:
    if partner == "Uber Eats":
        return [
            row["location_name"],
            row["partner_ref"] or "",
            row["channel_link_id"],
        ]
    return [row["location_name"], row["channel_link_id"]]


def partner_stores_csv_text(partner: str, rows: list[dict]) -> str:
    """CSV for one partner — same columns as previously listed in the email."""
    buf = StringIO()
    writer = csv.writer(buf)
    writer.writerow(_partner_csv_headers(partner))
    for row in rows:
        writer.writerow(_partner_csv_row(partner, row))
    return buf.getvalue()


def partner_stores_csv_filename(partner: str, cohort_tag: str) -> str:
    partner_slug = partner.lower().replace(" ", "_")
    return f"{_cohort_slug(cohort_tag)}_{partner_slug}_stores.csv"


def build_partner_email(
    partner: str,
    rows: list[dict],
    *,
    cohort_tag: str,
    action_date: date,
    go_live_date: Optional[date] = None,
    contact_name: str = "",
    contact_email: str = "",
    extra_notes: str = "",
) -> dict[str, str]:
    """Return subject + plain-text body for one delivery partner."""
    action_str = action_date.strftime("%d %B %Y")
    go_live_str = (go_live_date or action_date).strftime("%d %B %Y")
    cohort = cohort_tag or "this batch"
    sign_off = contact_name.strip() or "Deliverect / One Stop team"
    contact_line = f"\nFor questions, contact {contact_email}.\n" if contact_email else ""

    templates = PARTNER_EMAIL_TEMPLATES[partner]
    fill = {"cohort": cohort, "action_date": action_str, "go_live_date": go_live_str}

    subject = _fill_template(templates["subject"], **fill)
    intro = _fill_template(templates["intro"], **fill)
    steps = _fill_template(templates["steps"], **fill)

    store_count = len(rows)
    if store_count:
        stores_section = (
            f"This batch includes {store_count} store(s). "
            "Please see the attached CSV for the store list and IDs.\n"
        )
    else:
        stores_section = "There are no stores in this batch.\n"

    body_parts = [
        f"Hi {partner} team,\n\n",
        intro + "\n\n",
        "Please follow these steps:\n\n",
        steps + "\n\n",
        stores_section,
    ]

    if extra_notes.strip():
        body_parts.append(f"\nAdditional notes:\n{extra_notes.strip()}\n")

    body_parts.append(f"\nKind regards,\n{sign_off}\n")
    body_parts.append(contact_line)

    return {
        "partner": partner,
        "subject": subject,
        "body": "".join(body_parts),
        "store_count": store_count,
    }


def build_all_partner_emails(
    grouped: dict[str, list[dict]],
    *,
    cohort_tag: str,
    action_date: date,
    go_live_date: Optional[date] = None,
    contact_name: str = "",
    contact_email: str = "",
    extra_notes: str = "",
) -> dict[str, dict[str, str]]:
    emails = {}
    for partner in PARTNER_ORDER:
        emails[partner] = build_partner_email(
            partner,
            grouped.get(partner) or [],
            cohort_tag=cohort_tag,
            action_date=action_date,
            go_live_date=go_live_date,
            contact_name=contact_name,
            contact_email=contact_email,
            extra_notes=extra_notes,
        )
    return emails
